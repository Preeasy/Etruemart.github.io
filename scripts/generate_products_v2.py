#!/usr/bin/env python3
"""
基于AI图像识别结果生成产品数据
结合图片特征和文件名分类来生成准确的产品信息
"""
import json
import os
import urllib.parse
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter

IMAGE_BASE_URL = "https://raw.githubusercontent.com/Preeasy/images/main/Images/"
RECOGNITION_RESULTS_PATH = "/tmp/image_recognition_results.json"

# 分类配置 - 基于文件名前缀
category_config = {
    "Fashion Jewelry": {
        "slug": "fashion-jewelry",
        "category_en": "Fashion Jewelry",
        "category_cn": "时尚首饰",
        "materials": ["Alloy", "Stainless Steel", "Brass", "Acrylic", "Crystal", "Pearl", "Rhinestone"],
        "plating": ["Gold Plated", "Silver Plated", "Rose Gold Plated", "Rhodium Plated"],
        "price_ranges": [(0.15, 0.85), (0.50, 1.50), (0.80, 2.00), (0.20, 0.90), (0.60, 1.80)],
        "moq_list": [12, 24, 36, 48, 60],
        "sku_prefix": "YW-FJ",
        "desc_template": "Premium quality {name} crafted with {material}. {plating} finish for lasting shine. Perfect for retail and wholesale. OEM/ODM welcome.",
        "desc_cn": "优质{name}，采用{material}制作，{plating}工艺，持久闪耀。适合零售和批发。欢迎OEM/ODM定制。",
        "product_names": [
            ("Gold Necklace", "金色项链"), ("Silver Necklace", "银色项链"), 
            ("Gold Earrings", "金色耳环"), ("Silver Earrings", "银色耳环"),
            ("Gold Bracelet", "金色手链"), ("Silver Bracelet", "银色手链"),
            ("Gold Ring", "金色戒指"), ("Silver Ring", "银色戒指"),
            ("Gold Pendant", "金色吊坠"), ("Silver Pendant", "银色吊坠"),
            ("Rose Gold Necklace", "玫瑰金项链"), ("Rose Gold Earrings", "玫瑰金耳环"),
            ("Crystal Necklace", "水晶项链"), ("Pearl Necklace", "珍珠项链"),
            ("Gold Chain", "金色链条"), ("Silver Chain", "银色链条"),
            ("Gold Hoop Earrings", "金色圆环耳环"), ("Silver Hoop Earrings", "银色圆环耳环"),
            ("Gold Stud Earrings", "金色耳钉"), ("Silver Stud Earrings", "银色耳钉"),
            ("Gold Jewelry Set", "金色首饰套装"), ("Silver Jewelry Set", "银色首饰套装"),
            ("Rhinestone Brooch", "水钻胸针"), ("Gold Brooch", "金色胸针"),
            ("Silver Anklet", "银色脚链"), ("Gold Anklet", "金色脚链"),
            ("Gold Bangle", "金色手镯"), ("Silver Bangle", "银色手镯"),
            ("Fashion Pendant", "时尚吊坠"), ("Fashion Necklace", "时尚项链"),
            ("Fashion Earrings", "时尚耳环"), ("Fashion Bracelet", "时尚手链"),
            ("Acrylic Necklace", "亚克力项链"), ("Acrylic Earrings", "亚克力耳环"),
            ("Brass Necklace", "黄铜项链"), ("Brass Earrings", "黄铜耳环"),
            ("Crystal Earrings", "水晶耳环"), ("Crystal Bracelet", "水晶手链"),
            ("Pearl Earrings", "珍珠耳环"), ("Pearl Bracelet", "珍珠手链"),
        ]
    },
    "Garment Accessories": {
        "slug": "garment-accessories",
        "category_en": "Garment Accessories",
        "category_cn": "服装辅料",
        "materials": ["Metal", "Alloy", "Brass", "Stainless Steel", "Plastic", "Resin"],
        "plating": ["Gold Plated", "Silver Plated", "Antique Bronze", "Gunmetal", "Nickel Free"],
        "price_ranges": [(0.05, 0.35), (0.10, 0.50), (0.20, 0.80), (0.08, 0.40), (0.15, 0.60)],
        "moq_list": [100, 200, 500, 1000, 2000],
        "sku_prefix": "YW-GA",
        "desc_template": "Durable {name} for clothing manufacturing and DIY projects. {material} construction with {plating} finish. Easy to install and long-lasting performance.",
        "desc_cn": "耐用{name}，适用于服装制造和DIY项目。{material}材质，{plating}工艺。安装简便，经久耐用。"
    },
    "Bag accessories": {
        "slug": "bags-accessories",
        "category_en": "Bag Accessories",
        "category_cn": "箱包配件",
        "materials": ["Alloy", "Metal", "Zinc Alloy", "Stainless Steel", "Leather", "Plastic"],
        "plating": ["Gold Plated", "Silver Plated", "Antique Bronze", "Gunmetal", "Rose Gold"],
        "price_ranges": [(0.15, 0.60), (0.30, 1.00), (0.40, 1.20), (0.20, 0.80), (0.50, 1.50)],
        "moq_list": [50, 100, 200, 500, 1000],
        "sku_prefix": "YW-BA",
        "desc_template": "Premium {name} for handbag manufacturing and repair. {material} construction with {plating} finish. Sturdy and reliable for long-term use.",
        "desc_cn": "优质{name}，适用于箱包制造和维修。{material}材质，{plating}工艺。坚固可靠，经久耐用。"
    },
    "Hair Accessories": {
        "slug": "hair-accessories",
        "category_en": "Hair Accessories",
        "category_cn": "发饰",
        "materials": ["Acrylic", "Crystal", "Pearl", "Fabric", "Metal", "Rhinestone", "Resin"],
        "plating": ["Gold Plated", "Silver Plated", "Rose Gold Plated", "Rhodium Plated"],
        "price_ranges": [(0.25, 0.90), (0.40, 1.20), (0.30, 1.00), (0.20, 0.80), (0.50, 1.50)],
        "moq_list": [24, 36, 48, 60, 100],
        "sku_prefix": "YW-HA",
        "desc_template": "Stylish {name} to elevate your hairstyle. Made with {material} and {plating} finish. Comfortable to wear and perfect for all occasions.",
        "desc_cn": "时尚{name}，提升发型魅力。采用{material}制作，{plating}工艺。佩戴舒适，适合各种场合。"
    },
    "Toys_Gift": {
        "slug": "toys-gift",
        "category_en": "Toys & Gift",
        "category_cn": "玩具礼品",
        "materials": ["Plush", "Plastic", "Silicone", "Fabric", "Wood", "Resin"],
        "plating": ["None", "Painted", "Printed"],
        "price_ranges": [(0.50, 2.00), (0.80, 3.00), (1.00, 4.00), (0.30, 1.50), (0.60, 2.50)],
        "moq_list": [24, 48, 60, 100, 200],
        "sku_prefix": "YW-TG",
        "desc_template": "Fun and engaging {name} made with {material}. Safe and durable for kids and adults. Great for gifts, parties, and promotions.",
        "desc_cn": "有趣{name}，采用{material}制作。安全耐用，适合儿童和成人。适用于礼品、派对和促销活动。"
    },
    "Home_Decor_Crafts": {
        "slug": "home-decor-crafts",
        "category_en": "Home Decor & Crafts",
        "category_cn": "家居装饰工艺品",
        "materials": ["Resin", "Acrylic", "Crystal", "Wood", "Metal", "Ceramic", "Glass", "Fabric"],
        "plating": ["Gold Plated", "Silver Plated", "Painted", "Glazed", "None"],
        "price_ranges": [(0.30, 1.20), (0.50, 2.00), (0.80, 3.00), (0.40, 1.50), (0.60, 2.50)],
        "moq_list": [24, 48, 60, 100, 200],
        "sku_prefix": "YW-HD",
        "desc_template": "Beautiful {name} for home decoration and craft projects. Made with {material} and finished with {plating}. Perfect for adding charm to any space.",
        "desc_cn": "精美{name}，适用于家居装饰和手工项目。采用{material}制作，{plating}工艺。为空间增添魅力。",
        "product_names": [
            ("Resin Figurine", "树脂摆件"), ("Acrylic Decoration", "亚克力装饰"),
            ("Crystal Ornament", "水晶装饰品"), ("Wooden Craft", "木质工艺品"),
            ("Metal Wall Art", "金属墙饰"), ("Ceramic Vase", "陶瓷花瓶"),
            ("Glass Candle Holder", "玻璃烛台"), ("Fabric Garland", "布艺花环"),
            ("Decorative Tray", "装饰托盘"), ("Table Centerpiece", "餐桌摆件"),
            ("Wall Hanging", "壁挂装饰"), ("Shelf Decor", "搁板装饰"),
            ("Decorative Bowl", "装饰碗"), ("Garden Decoration", "花园装饰"),
            ("Seasonal Wreath", "季节花环"), ("Photo Frame", "相框"),
            ("Jewelry Organizer", "首饰收纳"), ("Plant Pot", "花盆"),
            ("Lamp Shade", "灯罩"), ("Curtain Tieback", "窗帘绑带"),
            ("Door Wreath", "门饰花环"), ("Bookend", "书挡"),
            ("Wall Clock", "挂钟"), ("Mirror Frame", "镜框"),
            ("Decorative Pillow", "装饰抱枕"), ("Rug", "地毯"),
        ]
    },
}

# 编号图片(001.jpg等)的分类映射 - 根据图片识别结果
def get_numbered_image_category(filename, recognition_result):
    """根据识别结果确定编号图片的类别"""
    if recognition_result and recognition_result.get('category'):
        cat = recognition_result['category']
        if cat == 'jewelry':
            return 'Fashion Jewelry'
        elif cat == 'garment':
            return 'Garment Accessories'
        elif cat == 'bag':
            return 'Bag accessories'
        elif cat == 'hair':
            return 'Hair Accessories'
        elif cat == 'toy':
            return 'Toys_Gift'
    return None

def load_recognition_results():
    """加载识别结果"""
    if os.path.exists(RECOGNITION_RESULTS_PATH):
        with open(RECOGNITION_RESULTS_PATH, 'r', encoding='utf-8') as f:
            results = json.load(f)
        return {r['filename']: r for r in results}
    return {}

def get_image_files():
    """获取图片文件列表，按分类组织"""
    image_dir = "/tmp/images_repo/Images"
    if not os.path.exists(image_dir):
        print("Image repo not found. Please clone the repo first.")
        return {}
    
    files = sorted([f for f in os.listdir(image_dir) if f.lower().endswith(('.jpg', '.jpeg', '.png'))])
    
    # 跳过分类标题图和banner图
    category_headers = {f for f in files if f.startswith(('01-', '02-', '03-', '04-', '05-', '06-'))}
    banner_images = {f for f in files if f.startswith('banner-')}
    product_files = [f for f in files if f not in category_headers and f not in banner_images]
    
    # 加载识别结果
    recognition = load_recognition_results()
    
    categorized = {}
    
    # 1. 按文件名分类的图片
    for cat_name in category_config:
        categorized[cat_name] = []
    
    for filename in product_files:
        fname_lower = filename.lower()
        matched = False
        
        # 按文件名前缀匹配
        if 'fashion' in fname_lower or 'jewelry' in fname_lower:
            categorized['Fashion Jewelry'].append(filename)
            matched = True
        elif 'garment' in fname_lower:
            categorized['Garment Accessories'].append(filename)
            matched = True
        elif 'bag' in fname_lower:
            categorized['Bag accessories'].append(filename)
            matched = True
        elif 'hair' in fname_lower:
            categorized['Hair Accessories'].append(filename)
            matched = True
        elif 'toy' in fname_lower or 'gift' in fname_lower:
            categorized['Toys_Gift'].append(filename)
            matched = True
        elif 'home' in fname_lower or 'decor' in fname_lower or 'craft' in fname_lower:
            categorized['Home_Decor_Crafts'].append(filename)
            matched = True
        
        # 2. 编号图片通过识别结果分类
        if not matched:
            rec = recognition.get(filename)
            cat = get_numbered_image_category(filename, rec)
            if cat and cat in categorized:
                categorized[cat].append(filename)
            else:
                # 默认放到Fashion Jewelry
                categorized['Fashion Jewelry'].append(filename)
    
    return categorized

def generate_products():
    """生成产品数据"""
    image_files = get_image_files()
    recognition = load_recognition_results()
    
    products = []
    excel_data = []
    
    for cat_name, files in image_files.items():
        if cat_name not in category_config:
            continue
        if not files:
            continue
        
        config = category_config[cat_name]
        
        for i, filename in enumerate(files):
            # URL编码文件名
            encoded_filename = urllib.parse.quote(filename)
            main_image = f"{IMAGE_BASE_URL}{encoded_filename}"
            
            # 获取识别结果
            rec = recognition.get(filename, {})
            rec_name_en = rec.get('name_en', '')
            rec_name_cn = rec.get('name_cn', '')
            
            # 策略：对于Fashion Jewelry，优先使用图像识别结果（更准确）
            # 对于其他类目，如果识别结果太通用，则使用product_names列表
            if cat_name == 'Fashion Jewelry':
                if rec_name_en and rec_name_en != 'Fashion Jewelry':
                    name_en = rec_name_en
                    name_cn = rec_name_cn
                elif 'product_names' in config:
                    name_en, name_cn = config['product_names'][i % len(config['product_names'])]
                else:
                    name_en = config['category_en']
                    name_cn = config['category_cn']
            elif 'product_names' in config:
                name_en, name_cn = config['product_names'][i % len(config['product_names'])]
            else:
                name_en = rec_name_en or config['category_en']
                name_cn = rec_name_cn or config['category_cn']
            
            # 选择材料和工艺
            material = config['materials'][i % len(config['materials'])]
            plating = config['plating'][i % len(config['plating'])]
            
            # 选择价格和MOQ
            price_range = config['price_ranges'][i % len(config['price_ranges'])]
            moq = config['moq_list'][i % len(config['moq_list'])]
            
            # 生成描述
            description = config['desc_template'].format(
                name=name_en.lower(),
                material=material.lower(),
                plating=plating
            )
            
            # 生成SKU
            sku = f"{config['sku_prefix']}-{str(i+1).zfill(3)}"
            
            # 生成产品ID
            product_id = 1783332968000 + len(products)
            
            product = {
                "id": product_id,
                "image": main_image,
                "images": [main_image],
                "category": {
                    "name": config['category_en'],
                    "slug": config['slug']
                },
                "name": name_en,
                "nameCn": name_cn,
                "description": description,
                "sku": sku,
                "moq": moq,
                "priceMin": round(price_range[0], 2),
                "priceMax": round(price_range[1], 2),
                "material": material,
                "plating": plating,
                "stockStatus": "IN_STOCK",
                "seller": "Yiwu Yeatru trading company"
            }
            
            products.append(product)
            
            # Excel数据
            excel_data.append({
                "image_url": main_image,
                "sku": sku,
                "name_cn": name_cn,
                "name_en": name_en,
                "moq": moq,
                "price_min": round(price_range[0], 2),
                "price_max": round(price_range[1], 2),
                "category_l1": config['category_en'],
                "category_l2": config['category_cn']
            })
    
    return products, excel_data

def generate_excel(excel_data, output_path):
    """生成Excel文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Product List"
    
    from openpyxl.styles import Font
    # 标题行
    headers = ["Product Image", "SKU", "Product Name (CN)", "Product Name (EN)", "MOQ", "Price (Min)", "Price (Max)", "Category L1", "Category L2"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = Font(bold=True)
    
    # 设置列宽
    column_widths = [50, 15, 25, 30, 8, 12, 12, 20, 20]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # 数据行
    for row, data in enumerate(excel_data, 2):
        ws.cell(row=row, column=2, value=data["sku"])
        ws.cell(row=row, column=3, value=data["name_cn"])
        ws.cell(row=row, column=4, value=data["name_en"])
        ws.cell(row=row, column=5, value=data["moq"])
        ws.cell(row=row, column=6, value=data["price_min"])
        ws.cell(row=row, column=7, value=data["price_max"])
        ws.cell(row=row, column=8, value=data["category_l1"])
        ws.cell(row=row, column=9, value=data["category_l2"])
    
    wb.save(output_path)

if __name__ == '__main__':
    products, excel_data = generate_products()
    
    # 保存site-data.json
    site_data = {"products": products}
    with open('/workspace/site-data.json', 'w', encoding='utf-8') as f:
        json.dump(site_data, f, ensure_ascii=False, indent=2)
    print(f"Saved {len(products)} products to site-data.json")
    
    # 保存Excel
    excel_path = '/workspace/product_list.xlsx'
    generate_excel(excel_data, excel_path)
    print(f"Excel file saved to {excel_path}")
    
    # 复制到public目录
    import shutil
    shutil.copy(excel_path, '/workspace/public/product_list.xlsx')
    print("Excel file copied to public/product_list.xlsx")
    
    # 打印统计
    from collections import Counter
    cat_counter = Counter(p['category']['name'] for p in products)
    print("\nProducts by category:")
    for cat, count in cat_counter.most_common():
        print(f"  {cat}: {count}")
    
    # 打印样例
    print("\nSample products (first 10):")
    for p in products[:10]:
        print(f"  {p['sku']} - {p['name']} ({p['nameCn']}) - {p['image'].split('/')[-1]}")
