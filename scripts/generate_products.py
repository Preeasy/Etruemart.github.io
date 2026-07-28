import os
import json
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image as ExcelImage
import urllib.request

IMAGE_BASE_URL = "https://raw.githubusercontent.com/Preeasy/images/main/Images/"

category_config = {
    "Fashion Jewelry": {
        "slug": "fashion-jewelry",
        "names_en": [
            "Fashion Jewelry Set", "Elegant Jewelry Piece", "Trendy Fashion Accessory",
            "Statement Jewelry", "Classic Jewelry Design", "Modern Style Jewelry",
            "Dainty Jewelry Item", "Fashion Earring Set", "Stylish Necklace",
            "Designer Jewelry Collection", "Premium Jewelry Piece", "Fashion Bracelet",
            "Unique Jewelry Design", "Chic Fashion Accessory", "Luxury Style Jewelry",
            "Vintage Inspired Jewelry", "Minimalist Jewelry", "Boho Fashion Jewelry",
            "Glamorous Jewelry Set", "Everyday Wear Jewelry", "Party Wear Jewelry",
            "Casual Fashion Jewelry", "Sleek Metal Jewelry", "Beaded Jewelry Piece",
            "Crystal Jewelry Set", "Pearl Fashion Jewelry", "Gemstone Style Jewelry",
            "Gold Tone Jewelry", "Silver Finish Jewelry", "Rose Gold Jewelry",
            "Layered Jewelry Set", "Charm Jewelry Piece", "Pendant Style Jewelry",
            "Hoop Earring Set", "Stud Earring Pack", "Drop Earring Style",
            "Cuff Bracelet Design", "Bangle Jewelry Set", "Chain Necklace Style",
            "Choker Necklace Set", "Lariat Style Jewelry", "Y-Necklace Design",
            "Multi-strand Jewelry", "Statement Necklace", "Delicate Chain",
            "Bold Jewelry Piece", "Subtle Elegance Jewelry", "Trendsetting Piece",
            "Fashion Forward Jewelry", "Timeless Classic Jewelry", "Contemporary Design",
            "Artisan Crafted Jewelry", "Handmade Style Piece"
        ],
        "names_cn": [
            "时尚首饰套装", "优雅珠宝单品", "潮流时尚配饰",
            "夸张首饰", "经典珠宝设计", "现代风格饰品",
            "精致珠宝单品", "时尚耳环套装", "时髦项链",
            "设计师珠宝系列", "高级珠宝单品", "时尚手链",
            "独特珠宝设计", "别致时尚配饰", "奢华风格珠宝",
            "复古风珠宝", "极简风饰品", "波西米亚风珠宝",
            "华丽珠宝套装", "日常佩戴饰品", "派对珠宝",
            "休闲时尚珠宝", "简约金属饰品", "串珠珠宝单品",
            "水晶珠宝套装", "珍珠时尚珠宝", "宝石风格饰品",
            "金色调珠宝", "银色饰面饰品", "玫瑰金珠宝",
            "多层珠宝套装", "魅力珠宝单品", "吊坠风格饰品",
            "圆环耳环套装", "耳钉组合装", "垂坠耳环款式",
            "开口手镯设计", "手镯珠宝套装", "链条项链款式",
            "项圈项链套装", "套索风格珠宝", "Y型项链设计",
            "多股珠宝", "夸张项链", "纤细链条",
            "醒目珠宝单品", "低调优雅饰品", "潮流单品",
            "前卫时尚珠宝", "永恒经典饰品", "现代设计款",
            "手工匠艺珠宝", "手工风格单品"
        ],
        "materials": ["Alloy", "Stainless Steel", "Brass", "Acrylic", "Crystal", "Pearl", "Rhinestone"],
        "plating": ["Gold Plated", "Silver Plated", "Rose Gold Plated", "Rhodium Plated"],
        "price_ranges": [(0.15, 0.85), (0.50, 1.50), (0.80, 2.00), (0.20, 0.90), (0.60, 1.80)],
        "moq_list": [12, 24, 36, 48, 60]
    },
    "Garment Accessories": {
        "slug": "garment-accessories",
        "names_en": [
            "Metal Snap Button", "Decorative Buckle", "Zipper Pull",
            "Rhinestone Brooch", "Fabric Patch", "Lace Trim",
            "Dress Clip", "Bow Tie", "Silk Ribbon",
            "Fur Pom Pom", "Bead Embellishment", "Sequin Applique",
            "Collar Stay", "Hook & Eye", "Snap Fastener",
            "Button Cover", "Shank Button", "Velcro Strap"
        ],
        "names_cn": [
            "金属按扣", "装饰扣", "拉链头",
            "水钻胸针", "布贴", "蕾丝花边",
            "裙夹", "蝴蝶结", "丝绸缎带",
            "毛绒球", "珠子装饰", "亮片贴花",
            "领撑", "风纪扣", "按扣",
            "纽扣套", "柄扣", "魔术贴"
        ],
        "materials": ["Alloy", "Brass", "Fabric", "Rhinestone", "Plastic", "Metal"],
        "plating": ["Gold Plated", "Silver Plated", "Antique Bronze", "Gunmetal"],
        "price_ranges": [(0.05, 0.35), (0.10, 0.50), (0.08, 0.40), (0.20, 0.80)],
        "moq_list": [100, 200, 500]
    },
    "Hair Accessories": {
        "slug": "hair-accessories",
        "names_en": [
            "Sparkle Hair Clip", "Satin Scrunchie", "Floral Hair Pin",
            "Pearl Headband", "Bow Hair Clip", "Metal Hair Comb",
            "Velvet Headband", "Crystal Hair Barrette", "Butterfly Hair Clip",
            "Beaded Hair Pin", "Tassel Hair Tie", "Sequined Headband"
        ],
        "names_cn": [
            "闪亮发夹", "缎面发圈", "花卉发簪",
            "珍珠发箍", "蝴蝶结发夹", "金属发梳",
            "天鹅绒发箍", "水晶发夹", "蝴蝶发夹",
            "串珠发簪", "流苏发绳", "亮片发箍"
        ],
        "materials": ["Acrylic", "Crystal", "Pearl", "Fabric", "Metal", "Rhinestone"],
        "plating": ["Gold Plated", "Silver Plated", "Rose Gold Plated"],
        "price_ranges": [(0.25, 0.90), (0.40, 1.20), (0.30, 1.00)],
        "moq_list": [24, 36, 48]
    },
    "Bag accessories": {
        "slug": "bags-accessories",
        "names_en": [
            "Bag Charm", "Metal Zipper", "Leather Strap",
            "Bag Hook", "Keychain Pendant", "Bag Lock",
            "Diamond Rivet", "Bag Tag", "Adjustable Strap",
            "Magnetic Snap", "Bag Feet", "Decorative Stud",
            "Chain Strap", "Bag Handle", "Toggle Clasp",
            "D-Ring", "Swivel Hook", "Shoulder Pad"
        ],
        "names_cn": [
            "包挂件", "金属拉链", "皮革肩带",
            "挂包钩", "钥匙扣吊坠", "包锁",
            "钻石铆钉", "包牌", "可调节肩带",
            "磁吸扣", "包底钉", "装饰钉",
            "链条肩带", "包把手", "插扣",
            "D形环", "旋转钩", "肩垫"
        ],
        "materials": ["Alloy", "Leather", "Plastic", "Metal", "Rhinestone"],
        "plating": ["Gold Plated", "Silver Plated", "Antique Bronze", "Gunmetal"],
        "price_ranges": [(0.15, 0.60), (0.30, 1.00), (0.20, 0.80), (0.40, 1.20)],
        "moq_list": [50, 100, 200]
    },
    "Home_Decor_Crafts": {
        "slug": "home-decor-crafts",
        "names_en": [
            "Ceramic Vase", "Decorative Candle", "Wall Hanging",
            "Wooden Figurine", "Glass Bowl", "Metal Lantern",
            "Fabric Throw Pillow", "Artificial Flower", "Resin Figurine",
            "Marble Coaster", "Woven Basket", "Crystal Ornament",
            "Stone Paperweight", "Brass Candlestick", "Porcelain Figurine",
            "Macrame Wall Decor", "Terracotta Pot", "Glass Vase",
            "Metal Sculpture", "Silk Flower Arrangement", "Bamboo Decor",
            "Cotton Tapestry", "Felt Craft", "Clay Figurine"
        ],
        "names_cn": [
            "陶瓷花瓶", "装饰蜡烛", "壁挂",
            "木雕摆件", "玻璃碗", "金属灯笼",
            "布艺抱枕", "人造花", "树脂摆件",
            "大理石杯垫", "编织篮", "水晶饰品",
            "石头镇纸", "黄铜烛台", "瓷塑摆件",
            "绳编壁挂", "陶土花盆", "玻璃花瓶",
            "金属雕塑", "绢花插花", "竹制装饰",
            "棉质挂毯", "毛毡工艺品", "黏土摆件"
        ],
        "materials": ["Ceramic", "Wood", "Glass", "Metal", "Fabric", "Resin", "Stone", "Marble"],
        "plating": [],
        "price_ranges": [(0.80, 2.50), (1.50, 4.00), (0.50, 1.80), (2.00, 5.00)],
        "moq_list": [12, 24, 48]
    },
    "Toys_Gift": {
        "slug": "toys-gift",
        "names_en": [
            "Squishy Toy", "Fidget Spinner", "LED Light Up Toy",
            "Stuffed Animal", "Keychain Toy", "Novelty Pen",
            "Party Popper", "Glow Stick", "Finger Puppet",
            "Mini Puzzle", "Wind-up Toy", "Water Gun",
            "Bubble Wand", "Mask", "Toy Car",
            "Doll Accessory", "Play Dough", "Card Game",
            "Temporary Tattoo", "Whistle", "Yoyo",
            "Rubber Duck", "Sticker Pack", "Toy Sword",
            "Magic Trick", "Miniature Toy", "Plush Keychain",
            "Balloon", "Coloring Book", "Toy Camera",
            "Fake Mustache", "Party Hat", "Toy Train",
            "Ring Pop", "Mini Figure", "Stress Ball"
        ],
        "names_cn": [
            "捏捏乐", "指尖陀螺", "LED发光玩具",
            "毛绒玩具", "钥匙扣玩具", "新奇笔",
            "派对礼花", "荧光棒", "手指玩偶",
            "迷你拼图", "发条玩具", "水枪",
            "泡泡棒", "面具", "玩具车",
            "娃娃配件", "橡皮泥", "纸牌游戏",
            "临时纹身", "口哨", "悠悠球",
            "橡皮鸭", "贴纸包", "玩具剑",
            "魔术道具", "迷你玩具", "毛绒钥匙扣",
            "气球", "涂色书", "玩具相机",
            "假胡子", "派对帽", "玩具火车",
            "戒指糖", "迷你人偶", "减压球"
        ],
        "materials": ["Silicone", "Plastic", "Rubber", "Fabric", "Paper", "Wood"],
        "plating": [],
        "price_ranges": [(0.15, 0.70), (0.30, 1.00), (0.50, 1.50), (0.20, 0.80)],
        "moq_list": [24, 48, 100, 200]
    }
}

description_templates = {
    "Fashion Jewelry": "High-quality fashion jewelry made with premium materials. Perfect for daily wear or special occasions. Trendy design that complements any outfit.",
    "Garment Accessories": "Durable garment accessories for clothing manufacturing and DIY projects. Easy to install and long-lasting performance.",
    "Hair Accessories": "Stylish hair accessories to elevate your hairstyle. Comfortable to wear and perfect for all hair types. Ideal for parties and everyday looks.",
    "Bag accessories": "Premium bag accessories for handbag repair and customization. Sturdy construction for long-term use.",
    "Home_Decor_Crafts": "Beautiful home decor crafts to enhance your living space. Unique designs that add charm to any room. Perfect for gifts and personal use.",
    "Toys_Gift": "Fun and entertaining toys for kids and adults alike. Safe materials and durable construction. Great for party favors and gift giving."
}

def get_image_files():
    repo_path = "/tmp/images_repo/Images"
    if not os.path.exists(repo_path):
        print("Repo not found")
        return {}
    
    files = {}
    for filename in os.listdir(repo_path):
        if filename.endswith(".jpg") and not filename.startswith("banner") and not filename.startswith("0"):
            matched = False
            for cat_name in category_config.keys():
                if cat_name.lower().replace(" ", "-") in filename.lower() or cat_name.lower().replace("_", "-") in filename.lower():
                    if cat_name not in files:
                        files[cat_name] = []
                    files[cat_name].append(filename)
                    matched = True
                    break
            if not matched:
                if "uncategorized" not in files:
                    files["uncategorized"] = []
                files["uncategorized"].append(filename)
    
    for cat in files:
        files[cat].sort()
    
    return files

def generate_products():
    image_files = get_image_files()
    products = []
    excel_data = []
    
    for cat_name, files in image_files.items():
        if cat_name not in category_config:
            continue
        
        config = category_config[cat_name]
        print(f"\nProcessing {cat_name}: {len(files)} images")
        
        for i, filename in enumerate(files):
            idx = i % len(config["names_en"])
            name_en = config["names_en"][idx]
            name_cn = config["names_cn"][idx]
            material = config["materials"][idx % len(config["materials"])]
            
            price_idx = idx % len(config["price_ranges"])
            price_min, price_max = config["price_ranges"][price_idx]
            
            moq = config["moq_list"][idx % len(config["moq_list"])]
            
            plating = ""
            if config["plating"]:
                plating = config["plating"][idx % len(config["plating"])]
            
            main_image = IMAGE_BASE_URL + filename.replace(" ", "%20")
            
            sku_prefix = config["slug"].replace("-", "")[:3].upper()
            sku = f"{sku_prefix}-{cat_name[:3].upper()}-{str(i+1).zfill(3)}"
            
            product = {
                "id": int(str(1783332968000 + len(products))),
                "image": main_image,
                "images": [main_image],
                "category": {
                    "name": cat_name.replace("_", " "),
                    "slug": config["slug"]
                },
                "name": name_en,
                "nameCn": name_cn,
                "description": description_templates[cat_name],
                "sku": sku,
                "moq": moq,
                "priceMin": round(price_min, 2),
                "priceMax": round(price_max, 2),
                "material": material,
                "plating": plating,
                "stockStatus": "IN_STOCK",
                "seller": "Yiwu Yeatru trading company"
            }
            
            products.append(product)
            
            excel_data.append([
                main_image,
                sku,
                name_cn,
                name_en,
                moq,
                f"${price_min:.2f} - ${price_max:.2f}",
                cat_name.replace("_", " "),
                name_en
            ])
    
    return products, excel_data

def save_site_data(products):
    site_data = {
        "siteName": "eTrue Mark",
        "products": products,
        "categories": [
            {"name": "Fashion Jewelry", "slug": "fashion-jewelry"},
            {"name": "Garment Accessories", "slug": "garment-accessories"},
            {"name": "Hair Accessories", "slug": "hair-accessories"},
            {"name": "Bags & Accessories", "slug": "bags-accessories"},
            {"name": "Home Decor & Crafts", "slug": "home-decor-crafts"},
            {"name": "Toys & Gift", "slug": "toys-gift"}
        ]
    }
    
    with open("/workspace/site-data.json", "w", encoding="utf-8") as f:
        json.dump(site_data, f, indent=2, ensure_ascii=False)
    
    print(f"\nSaved {len(products)} products to site-data.json")

def create_excel(excel_data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Product List"
    
    header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
    header_font = Font(color="ffffff", bold=True, size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = ["Product Image URL", "SKU", "Product Name (CN)", "Product Name (EN)", "MOQ", "Price", "Primary Category", "Secondary Category"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    for row_idx, row_data in enumerate(excel_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            if col_idx == 5:
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 25
    
    for row in range(1, len(excel_data) + 2):
        ws.row_dimensions[row].height = 25
    
    output_path = "/workspace/product_list.xlsx"
    wb.save(output_path)
    print(f"\nExcel file saved to {output_path}")
    
    public_path = "/workspace/public/product_list.xlsx"
    wb.save(public_path)
    print(f"Excel file copied to {public_path}")

if __name__ == "__main__":
    products, excel_data = generate_products()
    save_site_data(products)
    create_excel(excel_data)
